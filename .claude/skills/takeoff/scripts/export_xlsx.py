"""
Spreadsheet Export for Construction Takeoff Agent

Writes a list of LineItem objects to a formatted .xlsx workbook
organized by trade, with summary totals.
"""

from __future__ import annotations

import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.workbook.views import BookView
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

from models import LineItem


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_TRADE_FONT = Font(name="Calibri", bold=True, size=12, color="1F3864")
_TRADE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

_SUBTOTAL_FONT = Font(name="Calibri", bold=True, size=11)
_SUBTOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

_GRAND_TOTAL_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
_GRAND_TOTAL_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")

_CURRENCY_FMT = '"$"#,##0.00'
_NUMBER_FMT = '#,##0.00'
_PERCENT_FMT = '0.0%'
_THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

_INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light orange

# Manual-input columns (1-indexed): F=Unit Cost, I=Labor Rate, M=Unit Price
_INPUT_COLUMNS = {6, 9, 13}

_COLUMNS = [
    ("Category", 18),                   # A  (1)
    ("Description", 42),                # B  (2)
    ("Qty", 10),                        # C  (3)
    ("Sheets", 8),                      # D  (4)
    ("Unit", 8),                        # E  (5)
    ("Unit Cost", 12),                  # F  (6)  ← manual input
    ("Mat Total", 13),                  # G  (7)
    ("Mat %", 8),                       # H  (8)
    ("Labor Rate", 12),                 # I  (9)  ← manual input (%)
    ("Labor Total", 13),                # J  (10)
    ("Labor %", 8),                     # K  (11)
    ("L+M Cost", 13),                   # L  (12)
    ("Unit Price", 12),                 # M  (13) ← manual input
    ("Amount", 13),                     # N  (14)
    ("Gross Profit", 13),               # O  (15)
    ("GPM", 8),                         # P  (16)
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _apply_cell_style(cell, font=None, fill=None, alignment=None, fmt=None, border=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if fmt:
        cell.number_format = fmt
    if border:
        cell.border = border


def _write_header_row(ws, row: int, hide_sheets: bool = False):
    for col_idx, (title, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        _apply_cell_style(cell, _HEADER_FONT, _HEADER_FILL, _HEADER_ALIGN, border=_THIN_BORDER)
        if col_idx == 4 and hide_sheets:
            ws.column_dimensions[get_column_letter(col_idx)].width = 0
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_line_item(ws, row: int, item: LineItem):
    # A=Category, B=Description, C=Qty, D=Sheets, E=Unit, F=Unit Cost,
    # G=Material Total, H=Material %, I=Labor Rate (%), J=Labor Total,
    # K=Labor %, L=L+M Cost, M=Unit Price, N=Amount, O=Gross Profit, P=GPM
    r = row
    sheets_val = item.sheets if item.sheets > 0 else ""
    # Material Total: use Sheets×UnitCost when sheets present, else Qty×UnitCost
    mat_formula = f"=IF(D{r}>0,D{r}*F{r},C{r}*F{r})"
    values = [
        item.category,                          # A (1)
        item.description,                       # B (2)
        item.quantity,                          # C (3)
        sheets_val,                             # D (4) Sheets
        item.unit,                              # E (5)
        0,                                      # F (6) Unit Cost — manual input
        mat_formula,                            # G (7) Material Total
        f'=IF(N{r}=0,"",G{r}/N{r})',           # H (8) Material %
        0,                                      # I (9) Labor Rate — manual input (%)
        f"=I{r}*N{r}",                         # J (10) Labor Total = Labor Rate × Amount
        f'=IF(N{r}=0,"",J{r}/N{r})',           # K (11) Labor %
        f"=G{r}+J{r}",                         # L (12) Labor + Materials Cost
        0,                                      # M (13) Unit Price — manual input
        f"=IF(D{r}>0,D{r}*M{r},C{r}*M{r})",     # N (14) Amount (per sheet when sheets present)
        f"=N{r}-L{r}",                         # O (15) Gross Profit
        f'=IF(N{r}=0,"",O{r}/N{r})',           # P (16) GPM
    ]
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(vertical="center")
        # Orange fill for manual-input columns
        if col_idx in _INPUT_COLUMNS:
            cell.fill = _INPUT_FILL
        # Number formats
        if col_idx in (6, 7, 10, 12, 13, 14, 15):
            cell.number_format = _CURRENCY_FMT
        elif col_idx in (3, 4):
            cell.number_format = _NUMBER_FMT
        elif col_idx in (8, 9, 11, 16):
            cell.number_format = _PERCENT_FMT


def _write_subtotal_row(ws, row: int, trade: str, mat_total: float, lab_total: float,
                        grand: float, first_data_row: int = 0, last_data_row: int = 0,
                        has_sheets: bool = False):
    # Layout: G=MatTotal, J=LabTotal, L=L+MCost, N=Amount, O=GrossProfit
    cell = ws.cell(row=row, column=1, value=f"{trade} Subtotal")
    _apply_cell_style(cell, _SUBTOTAL_FONT, _SUBTOTAL_FILL, border=_THIN_BORDER)
    for col_idx in range(2, 7):
        c = ws.cell(row=row, column=col_idx)
        _apply_cell_style(c, fill=_SUBTOTAL_FILL, border=_THIN_BORDER)

    # Sheets total (column D=4) if trade has sheet items
    if has_sheets and first_data_row > 0 and last_data_row > 0:
        fr, lr = first_data_row, last_data_row
        c = ws.cell(row=row, column=4, value=f"=SUM(D{fr}:D{lr})")
        _apply_cell_style(c, _SUBTOTAL_FONT, _SUBTOTAL_FILL, fmt=_NUMBER_FMT, border=_THIN_BORDER)

    if first_data_row > 0 and last_data_row > 0:
        fr, lr = first_data_row, last_data_row
        formulas = [
            (7,  f"=SUM(G{fr}:G{lr})"),     # Material Total
            (10, f"=SUM(J{fr}:J{lr})"),     # Labor Total
            (12, f"=SUM(L{fr}:L{lr})"),     # Labor + Materials Cost
            (14, f"=SUM(N{fr}:N{lr})"),     # Amount
            (15, f"=SUM(O{fr}:O{lr})"),     # Gross Profit
        ]
    else:
        formulas = [
            (7, mat_total), (10, lab_total), (12, grand),
            (14, grand), (15, 0),
        ]
    for col_idx, val in formulas:
        c = ws.cell(row=row, column=col_idx, value=val)
        _apply_cell_style(c, _SUBTOTAL_FONT, _SUBTOTAL_FILL, fmt=_CURRENCY_FMT, border=_THIN_BORDER)

    # Material % = Material Total / Amount
    c = ws.cell(row=row, column=8, value=f'=IF(N{row}=0,"",G{row}/N{row})')
    _apply_cell_style(c, _SUBTOTAL_FONT, _SUBTOTAL_FILL, fmt=_PERCENT_FMT, border=_THIN_BORDER)

    # Labor % = Labor Total / Amount
    c = ws.cell(row=row, column=11, value=f'=IF(N{row}=0,"",J{row}/N{row})')
    _apply_cell_style(c, _SUBTOTAL_FONT, _SUBTOTAL_FILL, fmt=_PERCENT_FMT, border=_THIN_BORDER)

    # GPM = Gross Profit / Amount
    c = ws.cell(row=row, column=16, value=f'=IF(N{row}=0,"",O{row}/N{row})')
    _apply_cell_style(c, _SUBTOTAL_FONT, _SUBTOTAL_FILL, fmt=_PERCENT_FMT, border=_THIN_BORDER)

    # Fill remaining cells (no formula, just styled)
    for col_idx in (9, 13):
        c = ws.cell(row=row, column=col_idx)
        _apply_cell_style(c, fill=_SUBTOTAL_FILL, border=_THIN_BORDER)


# ---------------------------------------------------------------------------
# Main export
# ---------------------------------------------------------------------------

def export_estimate(
    line_items: list[LineItem],
    output_path: str,
    project_name: str = "",
    project_address: str = "",
    notes: list[tuple[str, list[str]]] | None = None,
    insulation_notes: list[tuple[str, list[str]]] | None = None,
    images: dict[str, str] | None = None,
    building_model: dict | None = None,
    code_notes: dict[str, list[tuple[str, list[str]]]] | None = None,
) -> str:
    """
    Export line items to a formatted .xlsx workbook.

    Args:
        line_items: All LineItem objects from all trade calculators.
        output_path: Path for the output .xlsx file.
        project_name: Optional project name for the title sheet.
        project_address: Optional address for the title sheet.
        notes: Optional list of (section_title, [bullet_items]) for property/trade sheets.
        insulation_notes: Optional separate notes for insulation sheet (includes code reqs).
        images: Optional dict of image paths {"street_view": "/path.jpg", "satellite": "/path.jpg"}.
        building_model: Optional BuildingModel dict for project description sheet (plans mode).
        code_notes: Optional trade-keyed building code notes for per-trade sheets.

    Returns:
        The output file path.
    """
    wb = Workbook()

    # Set window size so the file doesn't open stretched across the entire monitor
    # Values are in twips (1/20 of a point). ~15000x10000 ≈ a compact window.
    wb.views = [BookView(windowWidth=15000, windowHeight=10000)]

    # Remove the default blank sheet created by Workbook()
    wb.remove(wb.active)

    # ---- Project Description sheet (plans mode) ----
    if building_model:
        _build_project_description_sheet(wb, building_model, project_name, project_address)
    # ---- Property sheet with images and notes (address mode) ----
    elif images and any(images.values()):
        _build_property_sheet(wb, project_name, images, notes=notes)

    # ---- Per-trade sheets ----
    by_trade = _group_by_trade(line_items)
    for trade, items in by_trade.items():
        safe_name = trade.replace("/", "-")[:31]  # Excel sheet name limit
        ws_trade = wb.create_sheet(safe_name)
        # Determine trade-specific notes: code_notes override, then general notes
        trade_notes = (code_notes or {}).get(trade) or notes
        if trade == "insulation":
            _build_insulation_sheet(ws_trade, items, notes=insulation_notes or trade_notes)
        else:
            _build_trade_sheet(ws_trade, trade, items, notes=trade_notes)

    # ---- Detail sheet (last) ----
    ws_detail = wb.create_sheet("Detail")
    _build_detail_sheet(ws_detail, line_items, notes=notes)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


def _dim_to_feet(dim) -> float:
    """Convert a Dimension dict {feet, inches} to decimal feet."""
    if not dim or not isinstance(dim, dict):
        return 0.0
    return (dim.get("feet", 0) or 0) + ((dim.get("inches", 0) or 0) / 12.0)


def _dim_to_str(dim) -> str:
    """Convert a Dimension dict to a readable string like 12'-6\"."""
    if not dim or not isinstance(dim, dict):
        return "—"
    feet = dim.get("feet", 0) or 0
    inches = dim.get("inches", 0) or 0
    if feet == 0 and inches == 0:
        return "—"
    if inches == 0:
        return f"{feet}'-0\""
    return f"{feet}'-{inches}\""


def _clean_label(value: str) -> str:
    """Convert snake_case to Title Case."""
    if not value:
        return ""
    return value.replace("_", " ").title()


_SECTION_HEADER_FONT = Font(name="Calibri", bold=True, size=12, color="1F3864")
_SECTION_HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_TABLE_HEADER_FONT = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
_TABLE_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_ALT_ROW_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")


def _build_project_description_sheet(wb, building_model: dict,
                                     project_name: str = "", project_address: str = ""):
    """Create a 'Project' sheet with building overview and area measurements."""
    ws = wb.create_sheet("Project", 0)  # Insert as first sheet
    bm = building_model

    # Column widths
    for col, width in [("A", 22), ("B", 18), ("C", 16), ("D", 16), ("E", 14), ("F", 14), ("G", 14)]:
        ws.column_dimensions[col].width = width

    # ---- Title ----
    ws.merge_cells("A1:G1")
    title_cell = ws.cell(row=1, column=1, value=project_name or bm.get("project_name", "Project Description"))
    title_cell.font = Font(name="Calibri", bold=True, size=18, color="1F3864")
    title_cell.alignment = Alignment(horizontal="center")

    row = 3

    # ---- Project Info ----
    info_items = []
    addr = project_address or bm.get("project_address", "")
    if addr:
        info_items.append(("Address", addr))
    btype = bm.get("building_type", "")
    if btype:
        info_items.append(("Building Type", _clean_label(btype)))
    stories = bm.get("stories", 0)
    if stories:
        info_items.append(("Stories", str(stories)))
    sqft = bm.get("sqft", 0)
    if sqft:
        info_items.append(("Total Area", f"{int(sqft):,} SF"))
    cz = bm.get("climate_zone", "")
    if cz:
        edition = bm.get("iecc_code_edition", "")
        info_items.append(("Climate Zone", f"{cz}" + (f" (IECC {edition})" if edition else "")))

    for label, value in info_items:
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(name="Calibri", bold=True, size=10, color="333333")
        c = ws.cell(row=row, column=2, value=value)
        c.font = Font(name="Calibri", size=10, color="333333")
        row += 1

    row += 1

    # ---- Room Breakdown Table ----
    rooms = bm.get("rooms", [])
    if rooms:
        ws.merge_cells(f"A{row}:G{row}")
        c = ws.cell(row=row, column=1, value="Room Breakdown")
        c.font = _SECTION_HEADER_FONT
        c.fill = _SECTION_HEADER_FILL
        c.border = _THIN_BORDER
        for col_idx in range(2, 8):
            ws.cell(row=row, column=col_idx).fill = _SECTION_HEADER_FILL
            ws.cell(row=row, column=col_idx).border = _THIN_BORDER
        row += 1

        headers = ["Room", "Floor", "Length", "Width", "Area (SF)", "Ceiling Type", "Floor Finish"]
        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col_idx, value=h)
            c.font = _TABLE_HEADER_FONT
            c.fill = _TABLE_HEADER_FILL
            c.alignment = Alignment(horizontal="center")
            c.border = _THIN_BORDER
        row += 1

        total_room_area = 0
        for i, room in enumerate(rooms):
            l = _dim_to_feet(room.get("length"))
            w = _dim_to_feet(room.get("width"))
            area = round(l * w)
            total_room_area += area
            fill = _ALT_ROW_FILL if i % 2 == 0 else None
            vals = [
                room.get("name", f"Room {i+1}"),
                room.get("floor", 1),
                _dim_to_str(room.get("length")),
                _dim_to_str(room.get("width")),
                area if area > 0 else "",
                _clean_label(room.get("ceiling_type", "")) or "—",
                _clean_label(room.get("floor_finish", "")) or "—",
            ]
            for col_idx, val in enumerate(vals, start=1):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.border = _THIN_BORDER
                c.font = Font(name="Calibri", size=10)
                if col_idx == 5 and isinstance(val, (int, float)) and val > 0:
                    c.number_format = "#,##0"
                if fill:
                    c.fill = fill
            row += 1

        # Total row
        c = ws.cell(row=row, column=1, value="Total")
        c.font = Font(name="Calibri", bold=True, size=10)
        c.border = _THIN_BORDER
        for col_idx in range(2, 5):
            ws.cell(row=row, column=col_idx).border = _THIN_BORDER
        c = ws.cell(row=row, column=5, value=total_room_area)
        c.font = Font(name="Calibri", bold=True, size=10)
        c.number_format = "#,##0"
        c.border = _THIN_BORDER
        for col_idx in range(6, 8):
            ws.cell(row=row, column=col_idx).border = _THIN_BORDER
        row += 2

    # ---- Roof Summary ----
    roof = bm.get("roof", {})
    if roof:
        ws.merge_cells(f"A{row}:C{row}")
        c = ws.cell(row=row, column=1, value="Roof Summary")
        c.font = _SECTION_HEADER_FONT
        c.fill = _SECTION_HEADER_FILL
        c.border = _THIN_BORDER
        for col_idx in range(2, 4):
            ws.cell(row=row, column=col_idx).fill = _SECTION_HEADER_FILL
            ws.cell(row=row, column=col_idx).border = _THIN_BORDER
        row += 1
        roof_items = []
        if roof.get("style"):
            roof_items.append(("Style", _clean_label(roof["style"])))
        if roof.get("material"):
            roof_items.append(("Material", _clean_label(roof["material"])))
        if roof.get("pitch"):
            roof_items.append(("Pitch", f"{roof['pitch']}/12"))
        if roof.get("total_area_sf"):
            roof_items.append(("Total Area", f"{int(roof['total_area_sf']):,} SF"))
        sections = roof.get("sections", [])
        if sections:
            roof_items.append(("Sections", str(len(sections))))
        for label, value in roof_items:
            c = ws.cell(row=row, column=1, value=label)
            c.font = Font(name="Calibri", bold=True, size=10, color="333333")
            c.border = _THIN_BORDER
            c = ws.cell(row=row, column=2, value=value)
            c.font = Font(name="Calibri", size=10)
            c.border = _THIN_BORDER
            row += 1
        row += 1

    # ---- Foundation Summary ----
    foundation = bm.get("foundation", {})
    if foundation and foundation.get("type"):
        ws.merge_cells(f"A{row}:C{row}")
        c = ws.cell(row=row, column=1, value="Foundation Summary")
        c.font = _SECTION_HEADER_FONT
        c.fill = _SECTION_HEADER_FILL
        c.border = _THIN_BORDER
        for col_idx in range(2, 4):
            ws.cell(row=row, column=col_idx).fill = _SECTION_HEADER_FILL
            ws.cell(row=row, column=col_idx).border = _THIN_BORDER
        row += 1
        found_items = [("Type", _clean_label(foundation["type"]))]
        if foundation.get("area_sf"):
            found_items.append(("Area", f"{int(foundation['area_sf']):,} SF"))
        if foundation.get("perimeter_lf"):
            found_items.append(("Perimeter", f"{int(foundation['perimeter_lf']):,} LF"))
        for label, value in found_items:
            c = ws.cell(row=row, column=1, value=label)
            c.font = Font(name="Calibri", bold=True, size=10, color="333333")
            c.border = _THIN_BORDER
            c = ws.cell(row=row, column=2, value=value)
            c.font = Font(name="Calibri", size=10)
            c.border = _THIN_BORDER
            row += 1
        row += 1

    # ---- Wall Summary ----
    walls = bm.get("walls", [])
    if walls:
        ws.merge_cells(f"A{row}:C{row}")
        c = ws.cell(row=row, column=1, value="Wall Summary")
        c.font = _SECTION_HEADER_FONT
        c.fill = _SECTION_HEADER_FILL
        c.border = _THIN_BORDER
        for col_idx in range(2, 4):
            ws.cell(row=row, column=col_idx).fill = _SECTION_HEADER_FILL
            ws.cell(row=row, column=col_idx).border = _THIN_BORDER
        row += 1
        ext_walls = [w for w in walls if w.get("is_exterior")]
        int_walls = [w for w in walls if not w.get("is_exterior")]
        ext_lf = sum(_dim_to_feet(w.get("length")) for w in ext_walls)
        int_lf = sum(_dim_to_feet(w.get("length")) for w in int_walls)
        wall_items = [
            ("Exterior Walls", str(len(ext_walls))),
            ("Exterior LF", f"{round(ext_lf):,} LF"),
            ("Interior Walls", str(len(int_walls))),
            ("Interior LF", f"{round(int_lf):,} LF"),
        ]
        thicknesses = set(w.get("thickness", "") for w in walls if w.get("thickness"))
        if thicknesses:
            wall_items.append(("Framing", ", ".join(sorted(thicknesses))))
        for label, value in wall_items:
            c = ws.cell(row=row, column=1, value=label)
            c.font = Font(name="Calibri", bold=True, size=10, color="333333")
            c.border = _THIN_BORDER
            c = ws.cell(row=row, column=2, value=value)
            c.font = Font(name="Calibri", size=10)
            c.border = _THIN_BORDER
            row += 1

    # Zoom
    ws.sheet_view.zoomScale = 90


def _build_property_sheet(wb, project_name: str, images: dict[str, str],
                          notes: list[tuple[str, list[str]]] | None = None):
    """Create a 'Property' sheet with embedded images and notes."""
    try:
        from openpyxl.drawing.image import Image as XlImage
    except ImportError:
        XlImage = None

    ws = wb.create_sheet("Property", 0)  # Insert as first sheet

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value=project_name or "Property Overview")
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="1F3864")
    title_cell.alignment = Alignment(horizontal="center")

    row = 3

    # Embed images
    if XlImage:
        img_labels = [
            ("street_view", "Street View"),
            ("satellite", "Satellite / Aerial View"),
        ]
        for key, label in img_labels:
            path = images.get(key, "")
            if not path or not os.path.isfile(path):
                continue
            cell = ws.cell(row=row, column=1, value=label)
            cell.font = Font(name="Calibri", bold=True, size=12, color="1F3864")
            row += 1
            try:
                img = XlImage(path)
                max_w = 500
                if img.width > max_w:
                    ratio = max_w / img.width
                    img.width = max_w
                    img.height = int(img.height * ratio)
                ws.add_image(img, f"A{row}")
                rows_for_img = max(2, int(img.height / 18))
                row += rows_for_img + 1
            except Exception as e:
                ws.cell(row=row, column=1, value=f"(Image could not be embedded: {e})")
                row += 2

    # Notes sections below images
    if notes:
        row += 1
        _write_notes(ws, row - 2, notes)

    # Set column widths for readability
    ws.column_dimensions["A"].width = 50
    ws.sheet_properties.pageSetUpPr = None  # reset any page setup
    ws.sheet_view.zoomScale = 90


_NOTE_TITLE_FONT = Font(name="Calibri", bold=True, size=11, color="1F3864")
_NOTE_BODY_FONT = Font(name="Calibri", size=10, color="333333")


def _write_notes(ws, start_row: int, notes: list[tuple[str, list[str]]]) -> int:
    """Write notes sections below the estimate table. Returns next available row."""
    row = start_row + 2  # blank row gap
    for section_title, bullets in notes:
        cell = ws.cell(row=row, column=1, value=section_title)
        cell.font = _NOTE_TITLE_FONT
        row += 1
        for bullet in bullets:
            cell = ws.cell(row=row, column=1, value=f"  •  {bullet}")
            cell.font = _NOTE_BODY_FONT
            # Merge across columns for readability
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        row += 1  # blank row between sections
    return row


def _group_by_trade(items: list[LineItem]) -> dict[str, list[LineItem]]:
    groups: dict[str, list[LineItem]] = defaultdict(list)
    for item in items:
        groups[item.trade].append(item)
    return dict(groups)


def _build_summary_sheet(ws, items: list[LineItem], project_name: str, address: str):
    """High-level summary: one row per trade with totals."""
    # Title
    ws.merge_cells("A1:E1")
    title_cell = ws.cell(row=1, column=1, value="Construction Cost Estimate")
    _apply_cell_style(title_cell, Font(name="Calibri", bold=True, size=16, color="1F3864"))

    row = 2
    if project_name:
        ws.cell(row=row, column=1, value="Project:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=project_name)
        row += 1
    if address:
        ws.cell(row=row, column=1, value="Address:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=address)
        row += 1

    row += 1

    # Summary table header
    headers = ["Trade", "Material Cost", "Labor Cost", "Trade Total"]
    widths = [25, 18, 18, 18]
    for col_idx, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        _apply_cell_style(cell, _HEADER_FONT, _HEADER_FILL, _HEADER_ALIGN, border=_THIN_BORDER)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    row += 1
    by_trade = _group_by_trade(items)
    grand_mat = 0.0
    grand_lab = 0.0
    grand_total = 0.0

    for trade, trade_items in by_trade.items():
        mat = sum(i.material_total for i in trade_items)
        lab = sum(i.labor_total for i in trade_items)
        total = sum(i.line_total for i in trade_items)
        grand_mat += mat
        grand_lab += lab
        grand_total += total

        ws.cell(row=row, column=1, value=trade.title()).border = _THIN_BORDER
        for col_idx, val in [(2, mat), (3, lab), (4, total)]:
            c = ws.cell(row=row, column=col_idx, value=val)
            _apply_cell_style(c, fmt=_CURRENCY_FMT, border=_THIN_BORDER)
        row += 1

    # Grand total row
    ws.cell(row=row, column=1, value="GRAND TOTAL")
    _apply_cell_style(ws.cell(row=row, column=1), _GRAND_TOTAL_FONT, _GRAND_TOTAL_FILL, border=_THIN_BORDER)
    for col_idx, val in [(2, grand_mat), (3, grand_lab), (4, grand_total)]:
        c = ws.cell(row=row, column=col_idx, value=val)
        _apply_cell_style(c, _GRAND_TOTAL_FONT, _GRAND_TOTAL_FILL, fmt=_CURRENCY_FMT, border=_THIN_BORDER)


def _build_detail_sheet(ws, items: list[LineItem], notes: list[tuple[str, list[str]]] | None = None):
    """All line items grouped by trade on one sheet."""
    _write_header_row(ws, 1)
    row = 2
    by_trade = _group_by_trade(items)

    subtotal_rows = []  # track subtotal row numbers for grand total Amount formula

    for trade, trade_items in by_trade.items():
        # Trade header row
        cell = ws.cell(row=row, column=1, value=trade.upper())
        _apply_cell_style(cell, _TRADE_FONT, _TRADE_FILL, border=_THIN_BORDER)
        for col_idx in range(2, len(_COLUMNS) + 1):
            _apply_cell_style(ws.cell(row=row, column=col_idx), fill=_TRADE_FILL, border=_THIN_BORDER)
        row += 1

        first_data_row = row
        mat_total = 0.0
        lab_total = 0.0
        line_total = 0.0
        for item in trade_items:
            _write_line_item(ws, row, item)
            mat_total += item.material_total
            lab_total += item.labor_total
            line_total += item.line_total
            row += 1
        last_data_row = row - 1

        trade_has_sheets = any(getattr(i, "sheets", 0) > 0 for i in trade_items)
        _write_subtotal_row(ws, row, trade.title(), mat_total, lab_total, line_total,
                            first_data_row, last_data_row, has_sheets=trade_has_sheets)
        subtotal_rows.append(row)
        row += 2  # blank row between trades

    # Grand total
    grand_mat = sum(i.material_total for i in items)
    grand_lab = sum(i.labor_total for i in items)
    grand_total = sum(i.line_total for i in items)

    cell = ws.cell(row=row, column=1, value="GRAND TOTAL")
    _apply_cell_style(cell, _GRAND_TOTAL_FONT, _GRAND_TOTAL_FILL, border=_THIN_BORDER)
    for col_idx in range(2, 7):
        _apply_cell_style(ws.cell(row=row, column=col_idx), fill=_GRAND_TOTAL_FILL, border=_THIN_BORDER)
    # Grand total formulas — sum subtotal rows: G=MatTotal, J=LabTotal, L=L+MCost, N=Amount, O=GP
    for col_idx, col_letter in [(7, "G"), (10, "J"), (12, "L"), (14, "N"), (15, "O")]:
        refs = "+".join(f"{col_letter}{r}" for r in subtotal_rows)
        c = ws.cell(row=row, column=col_idx, value=f"={refs}")
        _apply_cell_style(c, _GRAND_TOTAL_FONT, _GRAND_TOTAL_FILL, fmt=_CURRENCY_FMT, border=_THIN_BORDER)
    # Material % grand total
    c = ws.cell(row=row, column=8, value=f'=IF(N{row}=0,"",G{row}/N{row})')
    _apply_cell_style(c, _GRAND_TOTAL_FONT, _GRAND_TOTAL_FILL, fmt=_PERCENT_FMT, border=_THIN_BORDER)
    # Labor % grand total
    c = ws.cell(row=row, column=11, value=f'=IF(N{row}=0,"",J{row}/N{row})')
    _apply_cell_style(c, _GRAND_TOTAL_FONT, _GRAND_TOTAL_FILL, fmt=_PERCENT_FMT, border=_THIN_BORDER)
    # GPM grand total
    c = ws.cell(row=row, column=16, value=f'=IF(N{row}=0,"",O{row}/N{row})')
    _apply_cell_style(c, _GRAND_TOTAL_FONT, _GRAND_TOTAL_FILL, fmt=_PERCENT_FMT, border=_THIN_BORDER)
    # Fill remaining cells
    for col_idx in (9, 13):
        _apply_cell_style(ws.cell(row=row, column=col_idx), fill=_GRAND_TOTAL_FILL, border=_THIN_BORDER)

    # Zoom
    ws.sheet_view.zoomScale = 90

    # Notes below the table
    if notes:
        _write_notes(ws, row, notes)


_CODE_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")  # light blue


def _build_insulation_sheet(ws, items: list[LineItem],
                            notes: list[tuple[str, list[str]]] | None = None):
    """Custom insulation sheet with a dedicated Code column for R-value requirements."""
    # Insulation-specific columns (simpler than full trade sheet)
    ins_cols = [
        ("Category", 20),       # A
        ("Description", 48),    # B
        ("Qty", 12),            # C
        ("Unit", 8),            # D
        ("Code", 28),           # E — building code R-value requirement
        ("Unit Cost", 12),      # F — manual input
        ("Material Total", 14), # G
        ("Unit Price", 12),     # H — manual input
        ("Amount", 14),         # I
    ]

    # Title
    ws.merge_cells("A1:E1")
    title = ws.cell(row=1, column=1, value="Insulation Estimate")
    _apply_cell_style(title, Font(name="Calibri", bold=True, size=14, color="1F3864"))

    # Header row
    row = 3
    for col_idx, (col_title, width) in enumerate(ins_cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_title)
        _apply_cell_style(cell, _HEADER_FONT, _HEADER_FILL, _HEADER_ALIGN, border=_THIN_BORDER)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    row = 4
    first_data_row = row
    for item in items:
        r = row
        values = [
            item.category,                          # A
            item.description,                       # B
            item.quantity,                           # C
            item.unit,                              # D
            getattr(item, "code_requirement", ""),   # E — Code
            0,                                      # F — Unit Cost (manual)
            f"=C{r}*F{r}",                         # G — Material Total
            0,                                      # H — Unit Price (manual)
            f"=C{r}*H{r}",                         # I — Amount
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            # Orange fill for manual-input columns
            if col_idx in (6, 8):
                cell.fill = _INPUT_FILL
            # Light blue fill for Code column
            if col_idx == 5:
                cell.fill = _CODE_FILL
                cell.font = Font(name="Calibri", bold=True, size=10, color="1F3864")
            # Number formats
            if col_idx == 3:
                cell.number_format = _NUMBER_FMT
            elif col_idx in (6, 7, 8, 9):
                cell.number_format = _CURRENCY_FMT
        row += 1

    last_data_row = row - 1

    # Subtotal row
    row += 1
    cell = ws.cell(row=row, column=1, value="Insulation Subtotal")
    _apply_cell_style(cell, _SUBTOTAL_FONT, _SUBTOTAL_FILL, border=_THIN_BORDER)
    for col_idx in range(2, len(ins_cols) + 1):
        _apply_cell_style(ws.cell(row=row, column=col_idx), fill=_SUBTOTAL_FILL, border=_THIN_BORDER)
    # Sum formulas
    for col_idx in (7, 9):  # G=Material Total, I=Amount
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=row, column=col_idx,
                       value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})")
        _apply_cell_style(cell, _SUBTOTAL_FONT, _SUBTOTAL_FILL, fmt=_CURRENCY_FMT, border=_THIN_BORDER)

    # Zoom
    ws.sheet_view.zoomScale = 90

    # Notes below the table
    if notes:
        _write_notes(ws, row, notes)


def _build_trade_sheet(ws, trade: str, items: list[LineItem],
                       notes: list[tuple[str, list[str]]] | None = None):
    """Individual trade sheet with just that trade's items."""
    # Detect if this trade has sheet items (drywall)
    has_sheets = any(getattr(item, "sheets", 0) > 0 for item in items)
    hide_sheets = not has_sheets

    # Trade title
    ws.merge_cells("A1:E1")
    title = ws.cell(row=1, column=1, value=f"{trade.title()} Estimate")
    _apply_cell_style(title, Font(name="Calibri", bold=True, size=14, color="1F3864"))

    _write_header_row(ws, 3, hide_sheets=hide_sheets)
    row = 4
    first_data_row = row

    mat_total = 0.0
    lab_total = 0.0
    line_total = 0.0
    for item in items:
        _write_line_item(ws, row, item)
        mat_total += item.material_total
        lab_total += item.labor_total
        line_total += item.line_total
        row += 1
    last_data_row = row - 1

    row += 1
    subtotal_row = row
    _write_subtotal_row(ws, row, trade.title(), mat_total, lab_total, line_total,
                        first_data_row, last_data_row, has_sheets=has_sheets)

    # Price per sheet summary row (only for trades with sheet items)
    if has_sheets:
        row += 1
        # D{subtotal_row} = total sheets, N{subtotal_row} = total amount
        c = ws.cell(row=row, column=1, value="Price Per Sheet")
        _apply_cell_style(c, _SUBTOTAL_FONT, border=_THIN_BORDER)
        for col_idx in range(2, len(_COLUMNS) + 1):
            _apply_cell_style(ws.cell(row=row, column=col_idx), border=_THIN_BORDER)
        # Formula: Total Amount / Total Sheets
        sr = subtotal_row
        c = ws.cell(row=row, column=13,
                    value=f'=IF(D{sr}=0,"",N{sr}/D{sr})')
        _apply_cell_style(c, _SUBTOTAL_FONT, fmt=_CURRENCY_FMT, border=_THIN_BORDER)

    # Zoom
    ws.sheet_view.zoomScale = 90

    # Notes below the table
    if notes:
        _write_notes(ws, row, notes)
